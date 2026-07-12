import gradio as gr
import re
import os
import html
import sqlite3
import logging
import tempfile
from datetime import datetime
from contextlib import contextmanager

from langdetect import detect, LangDetectException
import PyPDF2

try:
    import docx as docx_lib
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# -----------------------------
# LOGGING
# -----------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("email_ai_classifier")

# -----------------------------
# CARGA PEREZOSA DE MODELOS (lazy loading)
# -----------------------------

_summarizer = None
_embedding_model = None
_category_embeddings = None


def get_summarizer():
    global _summarizer
    if _summarizer is None:
        logger.info("Cargando modelo de resumen (summarization)...")
        from transformers import pipeline
        _summarizer = pipeline("summarization")
    return _summarizer


def get_embedding_model():
    global _embedding_model, _category_embeddings
    if _embedding_model is None:
        logger.info("Cargando modelo de embeddings (sentence-transformers)...")
        from sentence_transformers import SentenceTransformer
        _embedding_model = SentenceTransformer(
            "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2",
            device="cpu",
        )
        _category_embeddings = _embedding_model.encode(
            list(CATEGORIES.values()), convert_to_tensor=True
        )
    return _embedding_model, _category_embeddings


# -----------------------------
# CATEGORÍAS SEMÁNTICAS
# -----------------------------
CATEGORIES = {
    "Urgente": "email crítico que requiere acción inmediata o resolución rápida",
    "Importante pero no urgente": "email relevante de trabajo que necesita revisión posterior",
    "Spam/Promoción": "email publicitario, marketing o venta",
    "Informativo": "email informativo sin necesidad de acción inmediata",
}

# -----------------------------
# KEYWORDS INTELIGENTES
# -----------------------------
KEYWORDS = {
    "Urgente": ["urgente", "asap", "crítico", "error", "fallo"],
    "Spam/Promoción": ["oferta", "discount", "promo", "suscríbete"],
}
NEGATIONS = ["no", "not", "never", "sin", "jamás"]

# Palabras/frases típicas de phishing o urgencia artificial
PHISHING_KEYWORDS = [
    "verifica tu cuenta", "verify your account", "actualiza tus datos",
    "suspensión de cuenta", "account suspended", "haz clic aquí",
    "click here", "confirma tu contraseña", "premio", "has ganado",
    "you have won", "transferencia urgente", "wire transfer",
]

# Dominios de acortadores de URL habitualmente usados en phishing
SHORTENER_DOMAINS = [
    "bit.ly", "tinyurl.com", "goo.gl", "t.co", "ow.ly", "is.gd",
    "buff.ly", "rebrand.ly", "cutt.ly",
]

URL_REGEX = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)
DEADLINE_REGEX = re.compile(
    r"\b(hoy|mañana|today|tomorrow|antes de|before|deadline|fecha límite|24\s?h(oras)?|48\s?h(oras)?)\b",
    re.IGNORECASE,
)


def has_negation(text, keyword, window=4):
    words = re.sub(r"[.,;!?]", " ", text.lower()).split()
    for i, w in enumerate(words):
        if w == keyword.lower():
            start = max(0, i - window)
            context = words[start:i]
            if any(n in context for n in NEGATIONS):
                return True
    return False


# -----------------------------
# BASE DE DATOS (SQLite)
# -----------------------------
DB_PATH = "email_history.db"


@contextmanager
def get_db_connection():
    """Context manager que garantiza que la conexión se cierra siempre y que
    cada escritura es atómica (evita el problema de reescribir un JSON
    completo con múltiples usuarios/pestañas concurrentes)."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except sqlite3.Error:
        conn.rollback()
        raise
    finally:
        conn.close()


def init_db():
    with get_db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS emails (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                language TEXT,
                category TEXT,
                urgency_score INTEGER,
                summary TEXT,
                full_text TEXT,
                attachment INTEGER,
                explanation TEXT,
                phishing_flags TEXT
            )
            """
        )


def save_history(record):
    try:
        with get_db_connection() as conn:
            conn.execute(
                """
                INSERT INTO emails
                    (date, language, category, urgency_score, summary,
                     full_text, attachment, explanation, phishing_flags)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    record["date"],
                    record["language"],
                    record["category"],
                    record["urgency_score"],
                    record["summary"],
                    record["full_text"],
                    int(record["attachment"]),
                    record["explanation"],
                    record["phishing_flags"],
                ),
            )
    except sqlite3.Error as e:
        logger.error(f"Error guardando en la base de datos: {e}")


def fetch_history(limit=10):
    try:
        with get_db_connection() as conn:
            rows = conn.execute(
                "SELECT * FROM emails ORDER BY id DESC LIMIT ?", (limit,)
            ).fetchall()
            return rows
    except sqlite3.Error as e:
        logger.error(f"Error leyendo la base de datos: {e}")
        return []


def export_to_csv():
    import csv

    path = os.path.join(tempfile.gettempdir(), "email_history_export.csv")
    with get_db_connection() as conn:
        rows = conn.execute("SELECT * FROM emails ORDER BY id").fetchall()
        if not rows:
            return None
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, quoting=csv.QUOTE_ALL)
            writer.writerow(rows[0].keys())
            for r in rows:
                writer.writerow([r[k] for k in r.keys()])
    return path


def export_to_json():
    import json

    path = os.path.join(tempfile.gettempdir(), "email_history_export.json")
    with get_db_connection() as conn:
        rows = conn.execute("SELECT * FROM emails ORDER BY id").fetchall()
        if not rows:
            return None
        data = [dict(r) for r in rows]
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    return path


def export_to_excel():
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl no está instalado; no se puede exportar a Excel.")
        return None

    path = os.path.join(tempfile.gettempdir(), "email_history_export.xlsx")
    with get_db_connection() as conn:
        rows = conn.execute("SELECT * FROM emails ORDER BY id").fetchall()
        if not rows:
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial"

        headers = rows[0].keys()
        ws.append(list(headers))
        for r in rows:
            ws.append([r[k] for k in headers])

        # Autoajuste aproximado del ancho de columnas
        for i, col in enumerate(headers, start=1):
            max_len = max(
                [len(str(col))] + [min(len(str(r[col])), 80) for r in rows]
            )
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

        wb.save(path)
    return path


def download_history():
    csv_path = export_to_csv()
    json_path = export_to_json()
    xlsx_path = export_to_excel()
    return csv_path, json_path, xlsx_path


def clear_history():
    try:
        with get_db_connection() as conn:
            conn.execute("DELETE FROM emails")
        return "<p>🗑️ Historial borrado correctamente.</p>"
    except sqlite3.Error as e:
        logger.error(f"Error borrando el historial: {e}")
        return "<p>⚠️ No se pudo borrar el historial.</p>"


# -----------------------------
# EXTRACCIÓN DE ADJUNTOS (PDF, DOCX, TXT) — múltiples archivos
# -----------------------------
def extract_pdf_text(filepath):
    text = ""
    try:
        reader = PyPDF2.PdfReader(filepath)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    except (PyPDF2.errors.PdfReadError, OSError, ValueError) as e:
        logger.error(f"Error leyendo PDF '{filepath}': {e}")
    return text


def extract_docx_text(filepath):
    if not DOCX_AVAILABLE:
        logger.warning("python-docx no está instalado; no se puede leer .docx.")
        return ""
    text = ""
    try:
        doc = docx_lib.Document(filepath)
        text = "\n".join(p.text for p in doc.paragraphs)
    except (OSError, ValueError, KeyError) as e:
        logger.error(f"Error leyendo DOCX '{filepath}': {e}")
    return text


def extract_txt_text(filepath):
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except OSError as e:
        logger.error(f"Error leyendo TXT '{filepath}': {e}")
        return ""


def extract_attachments_text(files):
    """files: lista de objetos gr.File (o None). Soporta PDF, DOCX y TXT."""
    if not files:
        return ""
    combined = []
    for file in files:
        if file is None:
            continue
        path = file.name
        ext = os.path.splitext(path)[1].lower()
        if ext == ".pdf":
            combined.append(extract_pdf_text(path))
        elif ext == ".docx":
            combined.append(extract_docx_text(path))
        elif ext == ".txt":
            combined.append(extract_txt_text(path))
        else:
            logger.warning(f"Formato de adjunto no soportado: {ext}")
    return "\n".join(combined)


# -----------------------------
# DETECCIÓN DE IDIOMA (por párrafo)
# -----------------------------
def detect_language(text):
    """Detecta el idioma dominante analizando párrafo a párrafo, para no
    fallar cuando el email mezcla varios idiomas."""
    paragraphs = [p.strip() for p in text.split("\n") if p.strip()]
    if not paragraphs:
        return "unknown"

    counts = {}
    for p in paragraphs:
        if len(p) < 3:
            continue
        try:
            lang = detect(p)
            counts[lang] = counts.get(lang, 0) + 1
        except LangDetectException:
            continue

    if not counts:
        return "unknown"
    return max(counts, key=counts.get)


# -----------------------------
# CLASIFICACIÓN SEMÁNTICA
# -----------------------------
def semantic_classification(text):
    embedding_model, category_embeddings = get_embedding_model()
    from sentence_transformers import util

    text_embedding = embedding_model.encode(text, convert_to_tensor=True)
    scores = util.cos_sim(text_embedding, category_embeddings)[0]
    best_idx = scores.argmax().item()
    return list(CATEGORIES.keys())[best_idx], float(scores[best_idx])


# -----------------------------
# KEYWORD BOOST
# -----------------------------
def keyword_boost(text, category):
    text_lower = text.lower()
    for cat, words in KEYWORDS.items():
        for w in words:
            if w in text_lower and not has_negation(text_lower, w):
                return cat, f"Keyword detectada: '{w}'"
    return category, "Clasificación basada en embeddings semánticos"


# -----------------------------
# HEURÍSTICAS DE SPAM / PHISHING
# -----------------------------
def analyze_phishing_risk(text):
    flags = []
    text_lower = text.lower()

    # Frases típicas de phishing
    for phrase in PHISHING_KEYWORDS:
        if phrase in text_lower:
            flags.append(f"Frase sospechosa: '{phrase}'")

    # URLs acortadas
    urls = URL_REGEX.findall(text)
    for url in urls:
        for domain in SHORTENER_DOMAINS:
            if domain in url.lower():
                flags.append(f"URL acortada detectada: {url}")
                break

    # Exceso de mayúsculas (ruido tipo "¡GANA DINERO YA!")
    letters = [c for c in text if c.isalpha()]
    if letters:
        upper_ratio = sum(1 for c in letters if c.isupper()) / len(letters)
        if upper_ratio > 0.3 and len(letters) > 20:
            flags.append("Uso excesivo de mayúsculas")

    # Exceso de signos de exclamación
    if text.count("!") >= 3:
        flags.append("Uso excesivo de signos de exclamación")

    # Remitente vs dominio de los enlaces (heurística simple)
    sender_match = re.search(r"from:\s*.*@([\w\.-]+)", text_lower)
    if sender_match and urls:
        sender_domain = sender_match.group(1)
        for url in urls:
            link_domain_match = re.search(r"https?://([\w\.-]+)", url.lower())
            if link_domain_match:
                link_domain = link_domain_match.group(1)
                if sender_domain not in link_domain and link_domain not in sender_domain:
                    flags.append(
                        f"El dominio del remitente ({sender_domain}) no coincide "
                        f"con el del enlace ({link_domain})"
                    )
                    break

    return flags


# -----------------------------
# SCORE DE URGENCIA (0-100)
# -----------------------------
def compute_urgency_score(text, category, semantic_score):
    score = 0.0

    # Base según la categoría semántica
    category_base = {
        "Urgente": 55,
        "Importante pero no urgente": 30,
        "Spam/Promoción": 5,
        "Informativo": 10,
    }
    score += category_base.get(category, 10)

    # Confianza del modelo semántico (0-1) aporta hasta 20 puntos
    score += max(0.0, min(1.0, semantic_score)) * 20

    # Keywords de urgencia explícitas
    text_lower = text.lower()
    for w in KEYWORDS["Urgente"]:
        if w in text_lower and not has_negation(text_lower, w):
            score += 8

    # Menciones de fechas límite / plazos
    deadline_matches = DEADLINE_REGEX.findall(text)
    score += min(len(deadline_matches) * 5, 15)

    return int(max(0, min(100, round(score))))


# -----------------------------
# RESUMEN ROBUSTO
# -----------------------------
def safe_summary(text):
    if len(text.split()) < 40:
        return text[:150]
    try:
        summarizer = get_summarizer()
        return summarizer(text[:1000], max_length=80, min_length=25, do_sample=False)[0][
            "summary_text"
        ]
    except (ValueError, RuntimeError, IndexError, KeyError) as e:
        logger.error(f"Error generando el resumen: {e}")
        return text[:150]


# -----------------------------
# FUNCIÓN PRINCIPAL
# -----------------------------
def classify_email(email_text, attachments, progress=gr.Progress()):
    if not email_text or not email_text.strip():
        if not attachments:
            return "<p>⚠️ Escribe un email o adjunta un archivo para analizar.</p>"

    progress(0.05, desc="Extrayendo texto de los adjuntos...")
    attachment_text = extract_attachments_text(attachments)
    full_text = (email_text or "") + "\n" + attachment_text

    progress(0.2, desc="Detectando idioma...")
    language = detect_language(full_text)

    progress(0.35, desc="Clasificando el email (embeddings semánticos)...")
    category, semantic_score = semantic_classification(full_text)
    category, explanation = keyword_boost(full_text, category)

    progress(0.55, desc="Analizando riesgo de phishing / spam...")
    phishing_flags = analyze_phishing_risk(full_text)
    if phishing_flags:
        category = "Spam/Promoción" if not category == "Urgente" else category

    progress(0.8, desc="Calculando score de urgencia...")
    urgency_score = compute_urgency_score(full_text, category, semantic_score)

    progress(0.9, desc="Generando resumen...")
    summary = safe_summary(full_text)

    record = {
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "language": language,
        "category": category,
        "urgency_score": urgency_score,
        "summary": summary,
        "full_text": full_text,
        "attachment": bool(attachments),
        "explanation": explanation,
        "phishing_flags": "; ".join(phishing_flags),
    }
    save_history(record)
    progress(1.0, desc="Listo")

    # --- Construcción del HTML de salida, escapando SIEMPRE el texto de usuario ---
    safe_language = html.escape(language)
    safe_category = html.escape(category)
    safe_summary_text = html.escape(summary)
    safe_explanation = html.escape(explanation)

    phishing_html = ""
    if phishing_flags:
        items = "".join(f"<li>{html.escape(flag)}</li>" for flag in phishing_flags)
        phishing_html = f"""
        <h4>⚠️ Señales de posible spam/phishing</h4>
        <ul>{items}</ul>
        """

    urgency_color = "#e53935" if urgency_score >= 70 else "#fb8c00" if urgency_score >= 40 else "#43a047"

    html_output = f"""
    <div style='border:2px solid #4CAF50; padding:20px; border-radius:12px; background:#f9f9f9'>
        <h2>📧 Email AI Analysis</h2>
        <p><b>Idioma:</b> {safe_language}</p>
        <p><b>Categoría:</b> {safe_category}</p>
        <p><b>Score de urgencia:</b>
            <span style='color:{urgency_color}; font-weight:bold'>{urgency_score}/100</span>
        </p>

        <h4>Resumen</h4>
        <p>{safe_summary_text}</p>

        <h4>Explicación</h4>
        <p>{safe_explanation}</p>

        {phishing_html}
    </div>
    """
    return html_output


# -----------------------------
# HISTORIAL VISUAL CON CONTENIDO COMPLETO
# -----------------------------
def load_history():
    rows = fetch_history(limit=10)
    if not rows:
        return "<p>Sin historial todavía.</p>"

    html_parts = ["<h3>📊 Historial de Emails (últimos 10)</h3>"]
    for idx, r in enumerate(rows):
        safe_date = html.escape(r["date"])
        safe_category = html.escape(r["category"] or "")
        safe_full_text = html.escape(r["full_text"] or "")
        urgency = r["urgency_score"] if r["urgency_score"] is not None else "-"

        html_parts.append(
            f"""
            <div style='border:1px solid gray;margin:5px;padding:10px'>
                <b>{safe_date} | {safe_category} | Urgencia: {urgency}/100</b>
                <button onclick="document.getElementById('full_{idx}').style.display='block'">
                    Ver contenido completo
                </button>
                <div id='full_{idx}' style='display:none; margin-top:10px; white-space: pre-wrap; border-top:1px solid #ccc; padding-top:5px;'>
                    {safe_full_text}
                </div>
            </div>
            """
        )
    return "".join(html_parts)


# -----------------------------
# INTERFAZ GRADIO
# -----------------------------
def reset_inputs():
    return "", None, ""


with gr.Blocks() as demo:
    gr.Markdown("# 🤖 AI Email Classifier")

    email_input = gr.Textbox(lines=10, label="Email")
    attachments_input = gr.File(
        label="Adjuntos (PDF, DOCX, TXT - opcional, varios permitidos)",
        file_count="multiple",
    )
    output = gr.HTML()

    with gr.Row():
        analyze_btn = gr.Button("Analizar Email", variant="primary")
        clear_btn = gr.Button("Limpiar")
        history_btn = gr.Button("Ver Historial")
        clear_history_btn = gr.Button("Borrar Historial")

    download_btn = gr.Button("Descargar Historial (CSV / JSON / Excel)")
    with gr.Row():
        csv_file = gr.File(label="CSV")
        json_file = gr.File(label="JSON")
        xlsx_file = gr.File(label="Excel")

    analyze_btn.click(
        classify_email, inputs=[email_input, attachments_input], outputs=output
    )
    clear_btn.click(reset_inputs, inputs=None, outputs=[email_input, attachments_input, output])
    history_btn.click(load_history, outputs=output)
    clear_history_btn.click(clear_history, outputs=output)
    download_btn.click(
        download_history, inputs=None, outputs=[csv_file, json_file, xlsx_file]
    )

    demo.load(lambda: init_db(), inputs=None, outputs=None)

if __name__ == "__main__":
    init_db()
    demo.launch()
